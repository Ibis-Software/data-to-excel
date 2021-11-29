"""This application is for parsing well formatted xml and json files and convert them into flattened Excel tables"""

import argparse
import enum
import json
import logging
import sys
import time
from typing import Any

import openpyxl
import pandas
import xmltodict
from cool import F
from openpyxl.utils import get_column_letter

logging.basicConfig(format="%(asctime)s | %(levelname)s | %(message)s", level=logging.INFO, datefmt="%Y-%m-%d %H:%M:%S")


class Naming(enum.Enum):
	"""This class is to enable a single choice in renaming fields are tables"""
	FIELDS = 1
	TABLES = 2


def append_to_excel(excel_path: str, data_frame: pandas.DataFrame, sheet_name: str):
	"""This method creates a new sheet for placing the table/dataframe in an existing workbook"""
	with pandas.ExcelWriter(excel_path, mode="a", engine="openpyxl") as excel_file:
		data_frame.to_excel(excel_file, sheet_name=sheet_name, startcol=2, startrow=0)


def convert_serialized_to_excel(input_file: str, output_file: str, is_json_file: bool):
	"""This function is main function to convert an input JSON or XML file to an Excel file with a sheet for every
	flattened table """
	if is_json_file:
		json_string = extract_json(input_file)
	else:
		json_string = convert_xml_to_json(input_file)
	json_df = pandas.json_normalize(json_string)
	create_workbook(output_file)
	tables = extract_dataframes(json_df)
	sorted_tables = sorted(tables.items())
	object_names = []
	iterations = len(sorted_tables)
	display_progress(iterations=iterations)
	for i, (name, table) in enumerate(sorted_tables, start=1):
		table_extended, sheet_name = fetch_proper_names(dataframe=table, sheet_name=name)
		append_to_excel(output_file, table_extended, sheet_name)
		object_names.append(name)
		display_progress(i, iterations)
	print("")
	format_excel(output_file, object_names)


def convert_xml_to_json(input_file: str) -> str:
	"""This function input a XML file and outputs a json string"""
	if input_file.split(".")[-1] == "xml":
		with open(input_file) as xml_file:
			data_dict = xmltodict.parse(xml_file.read())
			xml_file.close()
			return json.dumps(data_dict) | F(json.loads)
	else:
		logging.critical(f"This is no .xml file: {input_file}\nPlease try again.")
		sys.exit(0)


def flatten_from_nested(again, tables):
	"""This function flattens tables from the second or higher level of nesting"""
	loop_again = again
	new_tables = tables
	for name, table in new_tables.copy().items():
		for column, cells in table.iteritems():
			change_table = False
			for i, value in enumerate(cells.values, start=1):
				if isinstance(value, list):
					new_column = f"{name}.{column}{i}"
					new_df = pandas.json_normalize(value)
					new_tables[new_column] = new_df
					change_table = True
			if change_table:
				new_tables.pop(name)
				new_tables[name] = table.drop(column, axis=1)
				loop_again = True
	return loop_again, new_tables


def create_short_name(name: str) -> str:
	"""This function makes sure a table name is shorter than 32 characters to fit in an excel sheet label"""
	short_name = ""
	is_to_long = len(name) > 31
	if not is_to_long:
		return name
	names = name.split(".")
	for value in names[:len(names) - 1]:
		short_name = f"{short_name}{value[0:2]}{value[len(value) - 1]}."
	potential_short_name = f"{short_name}{names[len(names) - 1]}"
	return (
		potential_short_name if len(potential_short_name) <= 31 else
		f"{potential_short_name[:30]}{potential_short_name[len(potential_short_name) - 1]}"
	)


def create_workbook(output_file: str):
	"""This function creates a new workbook with one worksheet named temp"""
	workbook = openpyxl.Workbook()
	sheet = workbook.active
	sheet.title = "temp"
	workbook.save(filename=output_file)


def display_progress(i=0, iterations=None):
	"""This function creates a progress bar in the command line"""
	if iterations is None:
		iterations = []
	empty = ""
	print("progress: |%s%s|" % (empty.rjust(i, '-'), empty.rjust(iterations - i, ' ')), end="\r")


def extract_dataframes(dataframe):
	"""This function is to extract flattened dataframes from dataframes with nested data in multiple levels"""
	tables_head = flatten_first_level(dataframe)
	return flatten_other_levels(tables_head)


def extract_json(input_file: str) -> Any:
	"""This function is to extract the JSON string from a json file"""
	if input_file.split(".")[-1] == "json":
		with open(input_file, encoding="utf-8") as json_file:
			json_data = json.load(json_file)
		return json_data
	logging.critical(f"This is no .json file: {input_file}\nPlease try again.")
	sys.exit(0)


def fetch_proper_names(dataframe: pandas.DataFrame, sheet_name: str) -> (pandas.DataFrame, str):
	"""This function is for enriching the table fields and table name with full name information"""
	dictionary = get_dictionary(Naming.FIELDS)
	tables = get_dictionary(Naming.TABLES)
	new_sheet_name = tables.get(sheet_name) or create_short_name(sheet_name)

	new_df = dataframe
	for field, value in dataframe.iteritems():
		if field in dictionary:
			full_value = dictionary[field]
			new_name = f"{full_value} ({field})"
		else:
			new_name = field
		new_df = new_df.rename({field: new_name}, axis="columns")
	return new_df, new_sheet_name


def flatten_first_level(dataframe):
	"""This function is for flattening the first level of a json object"""
	columns_list = []
	new_tables = {}
	for field, value in dataframe.iteritems():
		if not isinstance(value.values[0], list):
			columns_list.append(value.name)
		else:
			new_tables[value.name] = pandas.json_normalize(value.values[0])
	new_tables["ROOT"] = dataframe[columns_list]
	return new_tables


def flatten_other_levels(tables):
	"""This function loops through tables to scan for nested tables and to add new tables for these"""
	loop_again = False
	while True:
		loop_again, new_tables = flatten_from_nested(loop_again, tables)
		if not loop_again:
			break

	return new_tables


def format_excel(output_file: str, object_names: []):
	"""Function to format Excel to display the tables in a good format"""
	excel = openpyxl.open(output_file)
	excel.remove(excel["temp"])
	sheets = excel.sheetnames
	format_sheets(excel, object_names, sheets)
	excel.save(output_file)
	excel.close()


def format_sheets(excel, object_names, sheets):
	"""Function to format Excel sheets"""
	for i, sheet in enumerate(sheets):
		active_sheet = excel[sheet]
		active_sheet.sheet_view.showGridLines = False
		active_sheet.freeze_panes = 'D2'
		active_sheet['A1'] = object_names[i]
		active_sheet['c1'] = "nr"
		format_sheet_columns(active_sheet)


def format_sheet_columns(active_sheet):
	"""Function to format columns and cells"""
	for column in active_sheet.columns:
		column_name = get_column_letter(column[0].column)
		maximum_value = 0
		for cell in active_sheet[column_name]:
			val_to_check = len(str(cell.value))
			if val_to_check > maximum_value:
				maximum_value = val_to_check
		active_sheet.column_dimensions[column_name].width = maximum_value + 5


def get_dictionary(choice: Naming):
	"""Function to get the abbreviations and corresponding full names"""
	if choice == choice.FIELDS:
		return {"afd": "afdrukken", "alg": "algemene begrotingsgegevens", "altcde": "de alternatieve code",
		        "bbd": "staart", "bdr": "bedrag", "bdrpcthvh": "bedrag percentage hoeveelheid",
		        "beschsts": "beschikbaarheidsstatus voor ib.nl", "bgharb": "doorgerekende arbeid",
		        "bghmta": "doorgerekend materiaal", "bghmte": "doorgerekend materieel",
		        "bghoda": "doorgerekende onderaanneming", "bghtot": "doorgerekend bedrag",
		        "bglarb": "opslag arbeid", "bglmta": "opslag materiaal", "bglmte": "opslag materieel",
		        "bgloda": "opslag voor onderaanneming", "bgltot": "opslag totaal", "bgrvltid": "valuta id",
		        "bloksts": "blokkeerstatus voor ib.nl", "bstcde": "bestekcodering", "btoarb": "bruto arbeid",
		        "btohvh": "bruto hoeveelheid", "btomta": "bruto materiaal", "btomte": "bruto materieel",
		        "btooda": "bruto onderaanneming", "btostr": "bruto staart", "btotot": "bruto totaal",
		        "btwarb": "btw arbeid", "btwmta": "btw materiaal", "btwmte": "btw materieel",
		        "btwoda": "btw voor onderaanneming", "bva": "begrotingsvaluta's", "bvaid": "basisvaluta id",
		        "bwcarb": "nacalculatiecode voor arbeid", "bwcmta": "nacalculatiecode voor materiaal",
		        "bwcmte": "nacalculatiecode voor materieel", "bwcoda": "nacalculatiecode voor onderaanneming",
		        "cat": "categorie", "cclcde": "de code van de calculatie regel", "cmt": "commentaren",
		        "cmtid": "commentaar id", "datum": "datum", "dla": "drieletterafkorting",
		        "docintid": "document id", "dri": "doorrekenindicatie", "dtm": "datum",
		        "eanartcde": "gtin code van het artikel", "egs": "eigenschappen", "egsid": "eigenschap id",
		        "elt": "elementen", "eltid": "element id", "enh": "eenheid",
		        "enhprs": "de handmatig ingevulde eenheidsprijs", "enhprsmta": "eenheidsprijs materiaal",
		        "enhprsmte": "eenheidsprijs materieel", "enhprsoda": "eenheidsprijs voor onderaanneming",
		        "facarb": "factor arbeid", "fachvh": "factor hoeveelheid", "facmta": "factor materiaal",
		        "facmte": "factor materieel", "facoda": "factor voor onderaanneming", "fml": "formule",
		        "freq": "frequentie", "fto": "foto", "gtl": "getal", "hvh": "hoeveelheid",
		        "ibcode": "unieke code voor ib.nl", "inbjt": "inschrijfbiljet", "invind": "invoegindicatie",
		        "kid": "koppelings-ID", "klm": "kolom", "klmid": "kolom id",
		        "kltkrtpct": "klantkortingspercentage", "koers": "koers", "krtgrpcde": "kortingsgroepcode",
		        "kst": "kosten", "ktp": "kostenposten", "ktpid": "kostenpost id",
		        "levartcode": "leverancier artikel code", "levartprdt": "leverancier artikelprijs datum",
		        "levbrmatpr": "leverancier bruto materiaalprijs", "levgtincde": "leverancier gtin code",
		        "levkrtpct": "leverancierskortingspercentage", "levnaam": "leverancier naam",
		        "lngtxt": "langtekst", "loccde": "locatie codering", "mdl": "meetstaat modellen",
		        "mdlid": "model id", "mid": "middelen", "midcde": "middelcode", "midid": "middel id",
		        "mki": "modelkolom id", "mmk": "meetstaatmodelkolommen", "mpt": "multipliciteit",
		        "msc": "meetstaatcellen", "msk": "meetstaatkolommen", "msr": "meetstaatrijen",
		        "mst": "meetstaten", "mstid": "meetstaat id", "mstkid": "meetstaatkoppeling id",
		        "mtnguid": "meting guid", "nme": "naam", "nr": "nummer", "ntoarb": "netto arbeid",
		        "ntobto": "is netto of bruto", "ntomta": "netto materiaal", "ntomte": "netto materieel",
		        "ntooda": "netto onderaanneming", "ntostr": "netto staart", "ntotot": "netto totaal",
		        "offertenaam": "offertenaam", "oid": "object id", "oms": "omschrijving",
		        "ondcde": "onderhoudscode", "opm": "opmaak", "plncde": "plan codering", "pom": "meetstaat naam",
		        "prdfact": "productie capaciteit", "prjid": "project id", "prtid": "parent id",
		        "pstaard": "aard van de (sub)bestekspost t.b.v. afrekening", "rgl": "regel",
		        "rglid": "regel id", "rglnr": "regelnummer", "rij": "rij", "rko": "reservekopie",
		        "scenario": "scenario", "sgk": "standaard gekoppelde kolom id", "sie": "sectie",
		        "sjb": "sjablonen", "sjbid": "sjabloon id", "snt": "sneltoets", "srt": "soort",
		        "stk": "stuurcode (totalen hiervan worden verzameld en bijgehouden)", "sts": "status",
		        "stt": "staart", "stu": "structuren", "tblnme": "tabelnaam", "tblsrt": "tabelsoort",
		        "teken": "teken", "tij": "tijd", "tijenh": "tijdseenheid", "tkn": "tekening",
		        "tlt": "toelichting", "totuur": "uren", "tpe": "type", "txt": "tekst", "ulb": "uurloonbedragen",
		        "ulc": "uurlooncomponenten", "ulncde": "uurlooncode",
		        "untnrmbb": "tijdnorm bestaande bouw (uneto)",
		        "untnrmbl": "tijdnorm bestaande bouw leeg (uneto)",
		        "untnrmne": "tijdnorm nieuwbouw eenmalig (uneto)",
		        "untnrmnr": "tijdnorm nieuwbouw repeterend (uneto)", "unttaakcde": "uneto-taakcode",
		        "url": "url", "usr": "user", "uur": "aantal uren", "uurnrm": "uurnorm",
		        "uurnrmtpe": "uurnormtype", "vbld": "bevat voorblad", "vlgnr": "volgnummer", "vlt": "valuta",
		        "vrs": "versie", "vzp": "verzamelpunten", "vzpid": "stuurcode", "wde": "waarde",
		        "wzgdtm": "wijzigingsdatum", "TradbegrotingIbis.bgr.dtm": "datum",
		        "TradbegrotingIbis.bgr.oms": "omschrijving", "TradbegrotingIbis.bgr.ntotot": "netto totaal",
		        "TradbegrotingIbis.bgr.btomta": "bruto materiaal", "TradbegrotingIbis.bgr.usr": "user",
		        "TradbegrotingIbis.bgr.prjid": "project id", "TradbegrotingIbis.bgr.freq": "frequentie",
		        "TradbegrotingIbis.bgr.nme": "naam", "TradbegrotingIbis.bgr.vlt": "valuta",
		        "TradbegrotingIbis.bgr.ntooda": "netto onderaanneming",
		        "TradbegrotingIbis.bgr.ntostr": "netto staart",
		        "TradbegrotingIbis.bgr.btooda": "bruto onderaanneming", "TradbegrotingIbis.bgr.vrs": "versie",
		        "TradbegrotingIbis.bgr.totuur": "uren", "TradbegrotingIbis.bgr.btomte": "bruto materieel",
		        "TradbegrotingIbis.bgr.bvaid": "basisvaluta id",
		        "TradbegrotingIbis.bgr.inbjt": "inschrijfbiljet",
		        "TradbegrotingIbis.bgr.btoarb": "bruto arbeid", "TradbegrotingIbis.bgr.btostr": "bruto staart",
		        "TradbegrotingIbis.bgr.btotot": "bruto totaal",
		        "TradbegrotingIbis.bgr.ntomta": "netto materiaal",
		        "TradbegrotingIbis.bgr.ntomte": "netto materieel",
		        "TradbegrotingIbis.bgr.vbld": "bevat voorblad", "TradbegrotingIbis.bgr.tpe": "type",
		        "TradbegrotingIbis.bgr.rko": "reservekopie", "TradbegrotingIbis.bgr.ntoarb": "netto arbeid",
		        "TradbegrotingIbis.bgr.invind": "invoegindicatie"}
	return {"TradbegrotingIbis.alg": "algemene begrotingsgegevens", "TradbegrotingIbis.bbd": "staart",
	        "TradbegrotingIbis.bva": "begrotingsvaluta's", "TradbegrotingIbis.cmt": "commentaren",
	        "TradbegrotingIbis.egs": "eigenschappen", "TradbegrotingIbis.elt": "elementen",
	        "TradbegrotingIbis.ktp": "kostenposten", "TradbegrotingIbis.mdl": "meetstaat modellen",
	        "TradbegrotingIbis.mid": "middelen", "TradbegrotingIbis.mmk": "meetstaatmodelkolommen",
	        "TradbegrotingIbis.msc": "meetstaatcellen", "TradbegrotingIbis.msk": "meetstaatkolommen",
	        "TradbegrotingIbis.msr": "meetstaatrijen", "TradbegrotingIbis.mst": "meetstaten",
	        "TradbegrotingIbis.sjb": "sjablonen", "TradbegrotingIbis.stu": "structuren",
	        "TradbegrotingIbis.ulb": "uurloonbedragen", "TradbegrotingIbis.ulc": "uurlooncomponenten",
	        "TradbegrotingIbis.vzp": "verzamelpunten"}


def parse_arguments() -> dict[str, Any]:
	"""Function for command line arguments to run the application"""
	argument_parser = argparse.ArgumentParser()
	argument_parser.add_argument(
		"-i", "--inputpath", required=True, help="Path to the input file"
	)
	argument_parser.add_argument(
		"-o", "--outputpath", required=True, help="Path to the output file",
	)
	argument_group = argument_parser.add_mutually_exclusive_group(required=True)
	argument_group.add_argument("-j", "--json", action="store_true", help="Input file is a .json file")
	argument_group.add_argument("-x", "--xml", action="store_true", help="Input file is a .xml file")
	return vars(argument_parser.parse_args())


if __name__ == '__main__':
	arguments = parse_arguments()
	input_path = arguments["inputpath"]
	output_path = arguments["outputpath"]
	is_json = arguments["json"]

	start_time = time.perf_counter()
	try:
		convert_serialized_to_excel(input_path, output_path, is_json)
	except Exception as exception:
		logging.critical(f"This error happened: {exception.__str__()}\nPlease try again.")
		sys.exit(0)
	end_time = time.perf_counter()

	logging.info(f"Excel file with tables created in {end_time - start_time:0.4f} seconds: {output_path}")
